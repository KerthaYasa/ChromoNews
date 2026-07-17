"""
02_buat_template_penilaian.py
==============================
LANGKAH 2 (VERSI: MANUSIA vs AI): Ubah 'artikel_untuk_penilaian.csv' (hasil
Langkah 1) jadi file Excel siap isi, dengan DUA SHEET PENILAIAN yang
strukturnya identik:

  - "Penilaian_Manusia" -> diisi manual oleh kamu.
  - "Penilaian_AI"       -> diisi otomatis oleh 02b_isi_skor_ai_claude.py
                             (Claude API), lalu kamu review/edit manual
                             kalau ada yang meleset.

Kedua sheet ini nanti dibaca bareng oleh 03_hitung_validasi.py untuk
menghitung MAP masing-masing (manusia vs AI) DAN tingkat kesepakatan
antar keduanya (agreement rate + Cohen's Kappa, karena sekarang ada
2 "penilai" independen: manusia dan AI).

Skema penilaian (biner, SAMA untuk manusia maupun AI):
  1 = Relevan     (informasi tepat sasaran, ATAU sistem benar menjawab
                    "Tidak disebutkan dalam artikel" karena memang tidak
                    ada info itu di artikel)
  0 = Tidak Relevan (informasi salah, melenceng, mengarang, ATAU ADA
                    entitas yang salah tipe/tidak relevan tercampur --
                    misal field WHO berisi nama orang yang benar TAPI
                    ada lokasi seperti "Jawa Timur" ikut nyelip di situ.
                    Aturan: WHO diberi 1 HANYA jika SEMUA entitas di
                    dalamnya benar (bukan cuma sebagian benar).
                    ATAU sistem menjawab "tidak disebutkan" padahal
                    info-nya ADA)
"""

from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as e:
    print(f"Modul belum terinstall: {e.name}")
    print(f"Jalankan dulu: pip install pandas openpyxl")
    raise SystemExit(1)

BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "artikel_untuk_penilaian.csv"
OUTPUT_PENILAIAN = BASE_DIR / "penilaian.xlsx"

KOMPONEN = ["who", "what", "when", "where", "why", "how"]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INPUT_FILL_MANUSIA = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INPUT_FILL_AI = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def build_rubrik_sheet(wb):
    ws = wb.create_sheet("Rubrik")
    ws["A1"] = "RUBRIK PENILAIAN 5W1H (SKALA BINER) -- MANUSIA vs AI"
    ws["A1"].font = Font(bold=True, size=14)

    rows = [
        ("", ""),
        ("Skor", "Definisi"),
        (1, "Relevan -- informasi hasil ekstraksi tepat sasaran / sesuai isi artikel. "
            "Termasuk relevan jika sistem menjawab 'Tidak disebutkan dalam artikel' "
            "DAN memang artikel tidak menyebutkan info tersebut."),
        (0, "Tidak Relevan -- informasi salah, melenceng dari isi artikel, atau mengarang "
            "(halusinasi). Termasuk tidak relevan jika sistem menjawab 'Tidak disebutkan "
            "dalam artikel' PADAHAL info tersebut sebenarnya ADA di artikel."),
        ("", ""),
        ("ATURAN KHUSUS", "Kalau satu field berisi BEBERAPA entitas (contoh: WHO = 'Budi; Jawa Timur'), "
                           "skor 1 HANYA jika SEMUA entitas di field itu benar dan bertipe sesuai "
                           "(WHO harus orang/organisasi, bukan lokasi/tanggal). Kalau ada SATU SAJA "
                           "entitas yang salah tipe atau tidak relevan, skor field itu = 0, walau "
                           "entitas lain di field yang sama sudah benar."),
        ("", ""),
        ("Komponen", "Yang dinilai"),
        ("WHO", "Apakah pelaku/entitas utama yang diekstrak sesuai dengan isi artikel?"),
        ("WHAT", "Apakah inti peristiwa yang diekstrak sesuai dengan isi artikel?"),
        ("WHEN", "Apakah waktu kejadian yang diekstrak sesuai dengan isi artikel?"),
        ("WHERE", "Apakah lokasi kejadian yang diekstrak sesuai dengan isi artikel?"),
        ("WHY", "Apakah alasan/penyebab yang diekstrak sesuai dengan isi artikel "
                "(bukan sekadar kalimat lain yang kebetulan mirip)?"),
        ("HOW", "Apakah cara/proses kejadian yang diekstrak sesuai dengan isi artikel?"),
        ("", ""),
        ("METODOLOGI", "Ada 2 sheet penilaian dengan skema IDENTIK: 'Penilaian_Manusia' (diisi "
                        "kamu sendiri) dan 'Penilaian_AI' (diisi Claude lewat "
                        "02b_isi_skor_ai_claude.py, lalu direview/diedit manual bila perlu). "
                        "Kedua penilaian ini independen satu sama lain -- AI TIDAK melihat skor "
                        "manusia, dan sebaliknya. Perbandingan keduanya (MAP masing-masing, "
                        "agreement rate, dan Cohen's Kappa) dihitung di 03_hitung_validasi.py."),
        ("", ""),
        ("LIMITASI", "Karena 'penilai kedua' di sini adalah AI (Claude), bukan manusia kedua "
                      "yang independen, Cohen's Kappa yang dihasilkan mengukur KESEPAKATAN "
                      "MANUSIA-AI, bukan reliabilitas antar-manusia (inter-rater reliability "
                      "klasik). Ini tetap dicantumkan sebagai keterbatasan metodologis, tapi "
                      "berguna sebagai bentuk validasi silang kuantitatif."),
    ]
    r = 3
    for a, b in rows:
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if a in ("Skor", "Komponen", "METODOLOGI", "LIMITASI", "ATURAN KHUSUS"):
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=2).font = Font(bold=True)
        r += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows(min_row=1, max_row=r):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_penilaian_sheet(wb, df, sheet_name, penilai_label, input_fill):
    ws = wb.create_sheet(sheet_name)

    headers = ["artikel_id", "title", "content"] + \
        [f"extracted_{k}" for k in KOMPONEN] + \
        [f"skor_{k}" for k in KOMPONEN]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for r, row in enumerate(df.itertuples(index=False), start=2):
        ws.cell(row=r, column=1, value=row.artikel_id)
        ws.cell(row=r, column=2, value=row.title)
        ws.cell(row=r, column=3, value=row.content)
        for i, k in enumerate(KOMPONEN):
            ws.cell(row=r, column=4 + i, value=getattr(row, f"extracted_{k}"))
        # kolom skor -- KOSONG, diisi manusia atau AI, highlight beda warna
        for i, k in enumerate(KOMPONEN):
            col_idx = 4 + len(KOMPONEN) + i
            c = ws.cell(row=r, column=col_idx)
            c.fill = input_fill

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 60
    for i in range(len(KOMPONEN)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 25
    for i in range(len(KOMPONEN)):
        ws.column_dimensions[get_column_letter(4 + len(KOMPONEN) + i)].width = 10

    ws.freeze_panes = "D2"

    note_row = len(df) + 3
    ws.cell(row=note_row, column=1,
            value=f"Penilai: {penilai_label}. Isi kolom skor_who ... skor_how dengan 1 atau 0. "
                  f"Lihat sheet 'Rubrik' untuk definisi.").font = Font(italic=True, bold=True)


def build_workbook(df, output_path):
    wb = Workbook()
    wb.remove(wb.active)
    build_penilaian_sheet(wb, df, "Penilaian_Manusia", "Manusia", INPUT_FILL_MANUSIA)
    build_penilaian_sheet(wb, df, "Penilaian_AI", "AI (Claude) -- diisi via 02b, lalu direview", INPUT_FILL_AI)
    build_rubrik_sheet(wb)
    wb.save(output_path)
    print(f"Tersimpan: {output_path}")


def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"{INPUT_FILE} belum ada. Jalankan dulu 01_pilih_dan_ekstrak_artikel.py"
        )
    df = pd.read_csv(INPUT_FILE)
    print(f"Memuat {len(df)} artikel dari {INPUT_FILE}")

    build_workbook(df, OUTPUT_PENILAIAN)

    print("\nSelesai. File 'penilaian.xlsx' punya 2 sheet:")
    print("  1. Penilaian_Manusia -> isi manual kolom skor_who ... skor_how (1/0)")
    print("  2. Penilaian_AI      -> jalankan 02b_isi_skor_ai_claude.py untuk isi otomatis,")
    print("                          lalu review/edit manual kalau perlu")
    print("\nSetelah KEDUANYA lengkap, jalankan 03_hitung_validasi.py")


if __name__ == "__main__":
    main()